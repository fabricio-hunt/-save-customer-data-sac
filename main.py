"""
RPA - Sistema de SAC Bemol
===========================
Dependências: pip install pyautogui openpyxl pywin32 pyperclip
"""

import time
import re
import logging
import win32gui
import win32clipboard
import pyautogui
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─── CONFIGURAÇÕES ────────────────────────────────────────────────────────────
ARQUIVO_ENTRADA   = "Clientes_Fora_Padrao 5.xlsx"
ARQUIVO_SAIDA     = "Clientes_Extraidos_SAC.xlsx"
COLUNA_ID_CLIENTE = "ID_CLIENTE"

PAUSA_APOS_BUSCA  = 10.0
PAUSA_ENTRE_IDS   = 1.0

# ─── COORDENADAS DOS CAMPOS ───────────────────────────────────────────────────
COORDS = {
    "Cliente"    : (324, 74),
    "CPF"        : (449, 77),
    "Nome"       : (324, 98),
    "Telefone"   : (300, 120),
    "Limpar tudo": (648, 120),
}
# ──────────────────────────────────────────────────────────────────────────────

pyautogui.FAILSAFE = True
pyautogui.PAUSE    = 0.05

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("rpa_log.txt", encoding="utf-8"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)


# ─── CLIPBOARD ────────────────────────────────────────────────────────────────

def limpar_clipboard():
    try:
        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.CloseClipboard()
    except Exception:
        pass


def ler_clipboard() -> str:
    for _ in range(5):   # tenta até 5x aguardando o clipboard ser preenchido
        time.sleep(0.15)
        try:
            win32clipboard.OpenClipboard()
            try:
                if win32clipboard.IsClipboardFormatAvailable(win32clipboard.CF_UNICODETEXT):
                    data = win32clipboard.GetClipboardData(win32clipboard.CF_UNICODETEXT)
                    if data and data.strip():
                        return data.strip()
            finally:
                win32clipboard.CloseClipboard()
        except Exception:
            pass
    return ""


# ─── LER CAMPO ────────────────────────────────────────────────────────────────

def ler_campo(campo: str, pausa_extra: float = 0.0) -> str:
    """
    Clica no campo, Ctrl+A, Ctrl+C, lê clipboard.
    Garante foco na janela do SAC antes de copiar.
    """
    import win32gui, win32con, win32process
    x, y = COORDS[campo]

    # Traz a janela do SAC para frente antes de clicar
    hwnd = win32gui.WindowFromPoint((x, y))
    if hwnd:
        # Sobe até a janela de nível superior
        pai = hwnd
        while win32gui.GetParent(pai):
            pai = win32gui.GetParent(pai)
        try:
            win32gui.ShowWindow(pai, win32con.SW_RESTORE)
            win32gui.SetForegroundWindow(pai)
        except Exception:
            pass
    time.sleep(0.3)

    limpar_clipboard()

    pyautogui.click(x, y)
    time.sleep(0.5 + pausa_extra)

    pyautogui.hotkey("ctrl", "a")
    time.sleep(0.2)

    pyautogui.hotkey("ctrl", "c")

    texto = ler_clipboard()
    log.debug(f"  ler_campo('{campo}') → '{texto}'")
    return texto


# ─── INTERAÇÃO COM A TELA ─────────────────────────────────────────────────────

def clicar(campo: str):
    x, y = COORDS[campo]
    pyautogui.click(x, y)
    time.sleep(0.2)


def limpar_tela():
    clicar("Limpar tudo")
    time.sleep(0.4)


def digitar_id(id_cliente: str):
    clicar("Cliente")
    time.sleep(0.1)
    pyautogui.hotkey("ctrl", "a")
    time.sleep(0.05)
    pyautogui.press("delete")
    time.sleep(0.05)
    pyautogui.typewrite(str(id_cliente), interval=0.04)
    pyautogui.press("enter")
    time.sleep(PAUSA_APOS_BUSCA)


def cliente_encontrado() -> bool:
    cpf  = ler_campo("CPF", pausa_extra=1.0)
    nome = ler_campo("Nome")
    return len(re.sub(r"\D", "", cpf)) >= 8 or len(nome) >= 3


def extrair_dados() -> dict:
    dados = {"CPF_CNPJ": "", "PRIMEIRO_NOME": "", "SOBRENOME": "", "CONTATO": ""}
    dados["CPF_CNPJ"] = formatar_cpf_cnpj(ler_campo("CPF", pausa_extra=1.0))
    nome = ler_campo("Nome")
    dados["PRIMEIRO_NOME"], dados["SOBRENOME"] = split_nome(nome)
    dados["CONTATO"] = ler_campo("Telefone")
    return dados


# ─── DEBUG ────────────────────────────────────────────────────────────────────

def debug_leitura():
    log.info("=== MODO DEBUG ===")
    log.info("Carregue um cliente na tela e aguarde 5 segundos...")
    time.sleep(5)
    for campo in ["Cliente", "CPF", "Nome", "Telefone"]:
        texto = ler_campo(campo)
        log.info(f"  {campo:<15} lido='{texto}'")


# ─── UTILITÁRIOS ──────────────────────────────────────────────────────────────

def formatar_cpf_cnpj(valor: str) -> str:
    n = re.sub(r"\D", "", valor)
    if len(n) == 11:
        return f"{n[:3]}.{n[3:6]}.{n[6:9]}-{n[9:]}"
    if len(n) == 14:
        return f"{n[:2]}.{n[2:5]}.{n[5:8]}/{n[8:12]}-{n[12:]}"
    return valor


def split_nome(nome: str):
    p = nome.strip().split()
    if not p:       return "", ""
    if len(p) == 1: return p[0], ""
    return p[0], " ".join(p[1:])


def ler_ids_planilha(caminho: str, coluna: str) -> list:
    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    cabecalho = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if coluna not in cabecalho:
        wb.close()
        raise ValueError(f"Coluna '{coluna}' não encontrada.")
    ci = cabecalho.index(coluna)
    ids_vistos, ids = set(), []
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = row[ci]
        if v is not None:
            s = str(v).strip().split(".")[0]
            if s and s not in ids_vistos:
                ids_vistos.add(s); ids.append(s)
    wb.close()
    return ids


def salvar_excel(resultados: list, caminho: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes Extraidos"
    cols = ["ID_CLIENTE", "CPF_CNPJ", "PRIMEIRO_NOME", "SOBRENOME", "CONTATO", "STATUS"]

    hf    = PatternFill("solid", start_color="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    for i, t in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=t)
        c.fill = hf; c.font = hfont
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    ok_f  = PatternFill("solid", start_color="E8F5E9")
    err_f = PatternFill("solid", start_color="FFEBEE")
    lf = Font(name="Arial", size=10)
    for ri, reg in enumerate(resultados, 2):
        fill = ok_f if reg.get("STATUS") == "OK" else err_f
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=ri, column=ci, value=reg.get(col, ""))
            c.font = lf; c.fill = fill

    for l, w in [("A",14),("B",20),("C",22),("D",28),("E",18),("F",14)]:
        ws.column_dimensions[l].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    wb.save(caminho)
    log.info(f"Planilha salva: {caminho}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    import sys
    if "--debug" in sys.argv:
        debug_leitura()
        return

    log.info("=== Iniciando RPA Bemol SAC ===")
    log.info("Para parar: mova o mouse para o CANTO SUPERIOR ESQUERDO.")
    log.info("Iniciando em 3 segundos — coloque foco na janela do SAC...")
    time.sleep(3)

    try:
        ids = ler_ids_planilha(ARQUIVO_ENTRADA, COLUNA_ID_CLIENTE)
        log.info(f"Planilha carregada: {len(ids)} IDs únicos")
    except Exception as e:
        log.error(str(e)); return

    resultados = []
    for i, id_cliente in enumerate(ids, 1):
        log.info(f"[{i}/{len(ids)}] ID: {id_cliente}")
        reg = {"ID_CLIENTE": id_cliente}
        try:
            limpar_tela()
            digitar_id(id_cliente)

            if not cliente_encontrado():
                raise ValueError("Cliente não encontrado")

            dados = extrair_dados()
            reg.update(dados)
            reg["STATUS"] = "OK"
            log.info(
                f"  ✓ {dados['PRIMEIRO_NOME']} {dados['SOBRENOME']} "
                f"| {dados['CPF_CNPJ']} | {dados['CONTATO']}"
            )

        except pyautogui.FailSafeException:
            log.warning("PARADO pelo usuário.")
            break
        except Exception as e:
            log.warning(f"  ✗ {e}")
            reg.update({"CPF_CNPJ":"","PRIMEIRO_NOME":"","SOBRENOME":"","CONTATO":""})
            reg["STATUS"] = f"ERRO: {e}"

        resultados.append(reg)
        time.sleep(PAUSA_ENTRE_IDS)

    salvar_excel(resultados, ARQUIVO_SAIDA)
    ok = sum(1 for r in resultados if r["STATUS"] == "OK")
    log.info(f"=== Concluído: {ok} OK | {len(resultados)-ok} erros ===")


if __name__ == "__main__":
    main()