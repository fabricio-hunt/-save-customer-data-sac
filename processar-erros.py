"""
RPA - Sistema de SAC Bemol
===========================
Dependências: pip install pyautogui openpyxl pywin32 pyperclip
"""

import time
import re
import logging
import ctypes
import ctypes.wintypes
import win32gui
import win32con
import win32clipboard
import pyautogui
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─── DPI AWARENESS ────────────────────────────────────────────────────────────
try:
    ctypes.windll.shcore.SetProcessDpiAwareness(2)
except Exception:
    try:
        ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass

# ─── CONFIGURAÇÕES ────────────────────────────────────────────────────────────
ARQUIVO_ENTRADA   = "reprocessar_erros.txt"
ARQUIVO_SAIDA     = "Clientes_Extraidos_SAC.xlsx"

PAUSA_APOS_BUSCA  = 10.0
PAUSA_ENTRE_IDS   = 1.0

# ─── COORDENADAS DOS CAMPOS ───────────────────────────────────────────────────
# Detectadas automaticamente via detectar_coords.py em 2026-03-04
COORDS = {
    "Cliente"    : (363, 77),    # Edit[0] - campo ID cliente
    "CPF"        : (488, 77),    # Edit[1] - campo CPF/CNPJ
    "Nome"       : (460, 100),   # Edit[3] - campo Nome completo
    "Telefone"   : (342, 123),   # Edit[4] - campo Telefone
    "Limpar tudo": (644, 123),   # Botao &Limpar tudo
}
# ──────────────────────────────────────────────────────────────────────────────

pyautogui.FAILSAFE = True
pyautogui.PAUSE    = 0.05

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("rpa_log.txt", encoding="utf-8"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

# ─── WIN32 CONSTANTES ─────────────────────────────────────────────────────────
WM_GETTEXT       = 0x000D
WM_GETTEXTLENGTH = 0x000E


# ─── LER TEXTO VIA HWND (WM_GETTEXT) ─────────────────────────────────────────

def _encontrar_janela_sac():
    """Encontra a janela principal do SAC."""
    resultado = [None]
    def callback(hwnd, _):
        if win32gui.IsWindowVisible(hwnd):
            titulo = win32gui.GetWindowText(hwnd)
            if titulo and ("benchimol" in titulo.lower() or "sistema de vendas" in titulo.lower()):
                resultado[0] = hwnd
    win32gui.EnumWindows(callback, None)
    return resultado[0]


def _mapear_edits(hwnd_sac):
    """
    Retorna dict com os hwnd dos campos Edit mapeados por nome lógico.
    Mapeia pela posição Y,X (mesma ordem usada no detectar_coords.py).
    """
    edits = []
    def cb(h, _):
        cls = win32gui.GetClassName(h)
        if "edit" in cls.lower():
            rect = win32gui.GetWindowRect(h)
            cx = (rect[0] + rect[2]) // 2
            cy = (rect[1] + rect[3]) // 2
            edits.append((h, cx, cy, rect))
    try:
        win32gui.EnumChildWindows(hwnd_sac, cb, None)
    except Exception:
        pass
    # Ordenar por Y, depois X (mesma ordem do detectar_coords.py)
    edits.sort(key=lambda e: (e[3][1], e[3][0]))
    return edits


def _encontrar_hwnd_campo(campo):
    """
    Encontra o hwnd do controle Edit mais próximo da coordenada do campo.
    """
    hwnd_sac = _encontrar_janela_sac()
    if not hwnd_sac:
        return None

    edits = _mapear_edits(hwnd_sac)
    if not edits:
        return None

    x_alvo, y_alvo = COORDS[campo]
    # Encontrar o edit mais próximo da coordenada
    melhor = None
    melhor_dist = float("inf")
    for h, cx, cy, rect in edits:
        dist = abs(cx - x_alvo) + abs(cy - y_alvo)
        if dist < melhor_dist:
            melhor_dist = dist
            melhor = h

    log.debug("  _encontrar_hwnd_campo('%s') -> hwnd=%s dist=%d" % (campo, melhor, melhor_dist))
    return melhor


def _ler_texto_wmgettext(hwnd):
    """Lê texto de um controle via WM_GETTEXT (funciona com TCompEdit Delphi)."""
    if not hwnd:
        return ""
    try:
        length = ctypes.windll.user32.SendMessageW(hwnd, WM_GETTEXTLENGTH, 0, 0)
        if length <= 0:
            return ""
        buf = ctypes.create_unicode_buffer(length + 1)
        ctypes.windll.user32.SendMessageW(hwnd, WM_GETTEXT, length + 1, buf)
        return buf.value.strip()
    except Exception as e:
        log.debug("  WM_GETTEXT falhou: %s" % e)
        return ""


# ─── CLIPBOARD (fallback) ────────────────────────────────────────────────────

def limpar_clipboard():
    try:
        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.CloseClipboard()
    except Exception:
        pass


def ler_clipboard() -> str:
    for _ in range(5):
        time.sleep(0.15)
        try:
            win32clipboard.OpenClipboard()
            try:
                if win32clipboard.IsClipboardFormatAvailable(win32clipboard.CF_UNICODETEXT):
                    data = win32clipboard.GetClipboardData(win32clipboard.CF_UNICODETEXT)
                    if data and data.strip():
                        return data.strip()
            finally:
                win32clipboard.CloseClipboard()
        except Exception:
            pass
    return ""


# ─── LER CAMPO ────────────────────────────────────────────────────────────────

def ler_campo(campo: str, pausa_extra: float = 0.0) -> str:
    """
    Lê o texto de um campo do SAC.
    Método 1: WM_GETTEXT direto no hwnd (mais confiável para Delphi TCompEdit)
    Método 2: Fallback com Ctrl+A, Ctrl+C via clipboard
    """
    # --- Método 1: WM_GETTEXT (direto, sem clipboard) ---
    hwnd_campo = _encontrar_hwnd_campo(campo)
    if hwnd_campo:
        texto = _ler_texto_wmgettext(hwnd_campo)
        if texto:
            log.debug("  ler_campo('%s') WM_GETTEXT -> '%s'" % (campo, texto))
            return texto

    # --- Método 2: Fallback clipboard (Ctrl+A, Ctrl+C) ---
    x, y = COORDS[campo]

    # Traz a janela do SAC para frente
    hwnd = win32gui.WindowFromPoint((x, y))
    if hwnd:
        pai = hwnd
        while win32gui.GetParent(pai):
            pai = win32gui.GetParent(pai)
        try:
            win32gui.ShowWindow(pai, win32con.SW_RESTORE)
            win32gui.SetForegroundWindow(pai)
        except Exception:
            pass
    time.sleep(0.3)

    limpar_clipboard()
    pyautogui.click(x, y)
    time.sleep(0.5 + pausa_extra)
    pyautogui.hotkey("ctrl", "a")
    time.sleep(0.2)
    pyautogui.hotkey("ctrl", "c")

    texto = ler_clipboard()
    log.debug("  ler_campo('%s') clipboard -> '%s'" % (campo, texto))
    return texto


# ─── INTERAÇÃO COM A TELA ─────────────────────────────────────────────────────

def clicar(campo: str):
    x, y = COORDS[campo]
    pyautogui.click(x, y)
    time.sleep(0.2)


def limpar_tela():
    clicar("Limpar tudo")
    time.sleep(0.4)


def digitar_id(id_cliente: str):
    clicar("Cliente")
    time.sleep(0.1)
    pyautogui.hotkey("ctrl", "a")
    time.sleep(0.05)
    pyautogui.press("delete")
    time.sleep(0.05)
    pyautogui.typewrite(str(id_cliente), interval=0.04)
    pyautogui.press("enter")
    time.sleep(PAUSA_APOS_BUSCA)


def cliente_encontrado() -> bool:
    cpf  = ler_campo("CPF", pausa_extra=1.0)
    nome = ler_campo("Nome")
    return len(re.sub(r"\D", "", cpf)) >= 8 or len(nome) >= 3


def extrair_dados() -> dict:
    dados = {"Nome": "", "Sobrenome": "", "CPF": "", "Telefone": ""}
    dados["CPF"] = formatar_cpf_cnpj(ler_campo("CPF", pausa_extra=1.0))
    nome = ler_campo("Nome")
    dados["Nome"], dados["Sobrenome"] = split_nome(nome)
    dados["Telefone"] = ler_campo("Telefone")
    return dados


# ─── DEBUG ────────────────────────────────────────────────────────────────────

def debug_leitura():
    log.info("=== MODO DEBUG ===")
    log.info("Carregue um cliente na tela e aguarde 5 segundos...")
    time.sleep(5)
    for campo in ["Cliente", "CPF", "Nome", "Telefone"]:
        texto = ler_campo(campo)
        log.info(f"  {campo:<15} lido='{texto}'")


# ─── UTILITÁRIOS ──────────────────────────────────────────────────────────────

def formatar_cpf_cnpj(valor: str) -> str:
    n = re.sub(r"\D", "", valor)
    if len(n) == 11:
        return f"{n[:3]}.{n[3:6]}.{n[6:9]}-{n[9:]}"
    if len(n) == 14:
        return f"{n[:2]}.{n[2:5]}.{n[5:8]}/{n[8:12]}-{n[12:]}"
    return valor


def split_nome(nome: str):
    p = nome.strip().split()
    if not p:       return "", ""
    if len(p) == 1: return p[0], ""
    return p[0], " ".join(p[1:])


def ler_ids_txt(caminho: str) -> list:
    if not __import__('os').path.exists(caminho):
        raise ValueError(f"Arquivo '{caminho}' não encontrado.")
    ids_vistos, ids = set(), []
    with open(caminho, 'r', encoding='utf-8') as f:
        for linha in f:
            v = linha.strip()
            if v and v not in ids_vistos:
                ids_vistos.add(v); ids.append(v)
    return ids


def ler_resultados_anteriores(caminho: str) -> list:
    import os
    if not os.path.exists(caminho):
        return []
        
    try:
        wb = load_workbook(caminho, read_only=True, data_only=True)
        ws = wb.active
        cabecalho = [str(c.value) if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        
        resultados = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            reg = {}
            for i, col in enumerate(cabecalho):
                if i < len(row):
                    reg[col] = row[i] if row[i] is not None else ""
            resultados.append(reg)
            
        wb.close()
        return resultados
    except Exception as e:
        log.warning(f"Não foi possível ler os resultados anteriores: {e}")
        return []


def salvar_excel(resultados: list, caminho: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes Extraidos"
    cols = ["ID", "Nome", "Sobrenome", "CPF", "Telefone", "Status"]

    hf    = PatternFill("solid", start_color="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    for i, t in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=t)
        c.fill = hf; c.font = hfont
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    ok_f  = PatternFill("solid", start_color="E8F5E9")
    err_f = PatternFill("solid", start_color="FFEBEE")
    lf = Font(name="Arial", size=10)
    for ri, reg in enumerate(resultados, 2):
        fill = ok_f if reg.get("Status") == "OK" else err_f
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=ri, column=ci, value=reg.get(col, ""))
            c.font = lf; c.fill = fill

    for l, w in [("A",14),("B",22),("C",28),("D",20),("E",18),("F",14)]:
        ws.column_dimensions[l].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    wb.save(caminho)
    log.info(f"Planilha salva: {caminho}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    import sys
    if "--debug" in sys.argv:
        debug_leitura()
        return

    log.info("=== Iniciando RPA Bemol SAC ===")
    log.info("Para parar: mova o mouse para o CANTO SUPERIOR ESQUERDO.")
    log.info("Iniciando em 3 segundos — coloque foco na janela do SAC...")
    time.sleep(3)

    try:
        ids_totais = ler_ids_txt(ARQUIVO_ENTRADA)
        log.info(f"Arquivo TXT carregado: {len(ids_totais)} IDs totais")
    except Exception as e:
        log.error(str(e)); return

    # Ler resultados anteriores e filtrar
    resultados_anteriores = ler_resultados_anteriores(ARQUIVO_SAIDA)
    
    # Mantém os resultados que deram OK para não perdê-los quando salvarmos de novo
    resultados = [r for r in resultados_anteriores if str(r.get("Status", "")).upper() == "OK"]
    ids_ok = {str(r.get("ID")).strip() for r in resultados}
    
    # Filtra para processar apenas os IDs que não estão na lista de OKs
    ids_para_processar = [id_cliente for id_cliente in ids_totais if str(id_cliente).strip() not in ids_ok]
    
    if len(ids_para_processar) == 0:
        log.info("Todos os clientes já foram processados com sucesso!")
        return
        
    log.info(f"Retomando processamento: {len(ids_ok)} já estão OK. {len(ids_para_processar)} restantes para processar (incluindo erros e pendentes).")

    try:
        for i, id_cliente in enumerate(ids_para_processar, 1):
            log.info(f"[{i}/{len(ids_para_processar)}] ID: {id_cliente}")
            reg = {"ID": id_cliente}
            try:
                limpar_tela()
                digitar_id(id_cliente)
    
                if not cliente_encontrado():
                    raise ValueError("Cliente não encontrado")
    
                dados = extrair_dados()
                reg.update(dados)
                reg["Status"] = "OK"
                log.info(
                    f"  ✓ {dados['Nome']} {dados['Sobrenome']} "
                    f"| {dados['CPF']} | {dados['Telefone']}"
                )
    
            except pyautogui.FailSafeException:
                log.warning("PARADO pelo usuário.")
                break
            except Exception as e:
                log.warning(f"  ✗ {e}")
                reg.update({"Nome":"","Sobrenome":"","CPF":"","Telefone":""})
                reg["Status"] = f"ERRO: {e}"
    
            resultados.append(reg)
            
            # Salvar a cada 50 clientes para evitar perda de dados
            if i % 50 == 0:
                salvar_excel(resultados, ARQUIVO_SAIDA)
                
            time.sleep(PAUSA_ENTRE_IDS)
            
    except KeyboardInterrupt:
        log.warning("PARADO pelo usuário no terminal (Ctrl+C). Salvando progresso...")

    salvar_excel(resultados, ARQUIVO_SAIDA)
    ok = sum(1 for r in resultados if r.get("Status") == "OK")
    log.info(f"=== Concluído: {ok} OK | {len(resultados)-ok} erros ===")


if __name__ == "__main__":
    main()